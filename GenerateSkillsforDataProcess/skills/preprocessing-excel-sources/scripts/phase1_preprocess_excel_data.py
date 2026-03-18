"""
脚本：phase1_preprocess_excel_data.py

将 data_prepare/source_file 下的 Excel 工作簿按 sheet 导出为 CSV，供后续表格检索或结构化使用。

处理逻辑：
- 输入：data_prepare/source_file 下所有支持后缀的 Excel 文件（.xlsx / .xlsm / .xltx / .xltm）。
- 使用 openpyxl 以只读、data_only 模式打开工作簿，逐 sheet 读取单元格并写入 CSV（UTF-8，逗号分隔）。
- 输出目录：data_prepare/table_cache。
- 命名规则：
  - 若工作簿仅有 1 个 sheet：输出 <excel_stem>.csv；
  - 若有多个 sheet：每个 sheet 输出 <excel_stem>__<sheet_name>.csv，其中 sheet 名会做安全化（仅保留字母数字与 -_）。
- 依赖：openpyxl（pip install openpyxl）。若未安装，初始化时抛 ImportError。
- 可通过 RUN_BATCH 或直接调用 ExcelPreprocessor.excel_to_csv 控制单文件或批量处理。
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import List
from common_schema import infer_layer_from_source_tree, normalize_path, now_iso


# ============================================================================
# 【配置区域】
# ============================================================================
import sys as _sys
_PROJECT_ROOT = Path(__file__).resolve().parents[3]
_CONFIGURING = _PROJECT_ROOT / "skills" / "configuring-pipeline"
for _p in (_CONFIGURING, _PROJECT_ROOT):
    if str(_p) not in _sys.path:
        _sys.path.insert(0, str(_p))

try:
    from pipeline_config_loader import load_config as _load_config
    _cfg = _load_config(ensure_dirs=True)
    EXCEL_DIR = Path(_cfg["intermediate"]["source_file"])
    CSV_OUTPUT_DIR = Path(_cfg["intermediate"]["table_cache"])
    DATA_PREPARE_DIR = EXCEL_DIR.parent
    PROCESSED_DATA_DIR = DATA_PREPARE_DIR / "processed_data"
except Exception:
    _cfg = {}
    DATA_PREPARE_DIR = Path(__file__).resolve().parents[1]
    PROCESSED_DATA_DIR = DATA_PREPARE_DIR / "processed_data"
    EXCEL_DIR = DATA_PREPARE_DIR / "source_file"
    CSV_OUTPUT_DIR = PROCESSED_DATA_DIR / "table_cache"

EXCEL_SOURCE_MAP_PATH = CSV_OUTPUT_DIR / "_excel_source_map.jsonl"
_excel = _cfg.get("excel") or {}
RUN_BATCH = bool(_excel.get("run_batch", True))
SUPPORTED_EXCEL_SUFFIXES = set(_excel.get("supported_suffixes", [".xlsx", ".xlsm", ".xltx", ".xltm"]))
# ============================================================================
# 【配置区域结束】
# ============================================================================


class ExcelPreprocessor:
    def __init__(
        self,
        excel_dir: Path = EXCEL_DIR,
        output_dir: Path = CSV_OUTPUT_DIR,
    ) -> None:
        self.excel_dir = Path(excel_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        try:
            from openpyxl import load_workbook  # type: ignore

            self._load_workbook = load_workbook
        except Exception as e:  # pragma: no cover
            raise ImportError(
                "缺少依赖 openpyxl，无法读取 .xlsx/.xlsm。请先执行：pip install openpyxl"
            ) from e

    def _sanitize_sheet_name(self, name: str) -> str:
        safe = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in (name or "sheet"))
        safe = safe.strip("_")
        return safe or "sheet"

    def _append_excel_source_map(
        self,
        *,
        csv_path: Path,
        excel_path: Path,
        sheet_name: str,
    ) -> None:
        """
        轻量来源映射：
        仅记录 CSV <- Excel 的来源关系与知识层，供 phase2_generate_excel_chunk 继承。

        不使用 sidecar，避免“只有 Excel 直转表格有 sidecar，而 PDF/MD 抽表格没有”的不一致。
        采用 jsonl 追加模式，phase2 读取时按 csv_name 取最后一条即可。
        """
        EXCEL_SOURCE_MAP_PATH.parent.mkdir(parents=True, exist_ok=True)
        record = {
            "csv_name": csv_path.name,
            "csv_path": normalize_path(csv_path),
            "origin_excel_path": normalize_path(excel_path),
            "source_name": excel_path.name,
            "sheet_name": sheet_name,
            "knowledge_layer": infer_layer_from_source_tree(excel_path, self.excel_dir),
            "emitted_at": now_iso(),
        }
        with EXCEL_SOURCE_MAP_PATH.open("a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False))
            f.write("\n")

    def excel_to_csv(self, excel_path: Path) -> List[Path]:
        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(excel_path)

        suffix = excel_path.suffix.lower()
        if suffix not in SUPPORTED_EXCEL_SUFFIXES:
            raise ValueError(
                f"不支持的 Excel 后缀：{suffix}。建议使用 .xlsx/.xlsm；当前支持：{sorted(SUPPORTED_EXCEL_SUFFIXES)}"
            )

        wb = self._load_workbook(filename=str(excel_path), data_only=True, read_only=True)
        sheets = wb.worksheets
        out_paths: List[Path] = []

        single_sheet = len(sheets) == 1
        for ws in sheets:
            if single_sheet:
                out_name = f"{excel_path.stem}.csv"
            else:
                out_name = f"{excel_path.stem}__{self._sanitize_sheet_name(ws.title)}.csv"
            out_path = self.output_dir / out_name

            with out_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])
            out_paths.append(out_path)
            self._append_excel_source_map(
                csv_path=out_path,
                excel_path=excel_path,
                sheet_name=ws.title,
            )

        return out_paths


if __name__ == "__main__":
    print("=" * 80)
    print("预处理脚本：phase1_preprocess_excel_data.py")
    print("=" * 80)
    print(f"EXCEL_DIR: {EXCEL_DIR}")
    print(f"CSV_OUTPUT_DIR: {CSV_OUTPUT_DIR}")
    print("=" * 80)

    if not RUN_BATCH:
        print("RUN_BATCH=False：脚本不会自动批处理，请自行在代码中调用 ExcelPreprocessor().excel_to_processed_json(...)")
        raise SystemExit(0)

    CSV_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    pre = ExcelPreprocessor()

    excel_files: List[Path] = []
    for suf in sorted(SUPPORTED_EXCEL_SUFFIXES):
        excel_files.extend(sorted(Path(EXCEL_DIR).rglob(f"*{suf}")))

    if not excel_files:
        print(f"未找到 Excel 文件：{EXCEL_DIR}（支持后缀：{sorted(SUPPORTED_EXCEL_SUFFIXES)}）")
        raise SystemExit(1)

    for i, xls_path in enumerate(excel_files, 1):
        print(f"\n{'*' * 80}")
        print(f"进度 {i}/{len(excel_files)}: {xls_path.name}")
        print(f"{'*' * 80}")
        out_files = pre.excel_to_csv(xls_path)
        print(f"输出 {len(out_files)} 个 CSV：")
        for p in out_files:
            print(f"  - {p}")

    print("\n全部 Excel -> CSV 转换完成。")
    try:
        from pipeline_run_logger import append_run_record
        append_run_record(
            step_id="phase1_excel",
            script="phase1_preprocess_excel_data.py",
            status="success",
            files_processed=len(excel_files),
            detail={"output_dir": str(CSV_OUTPUT_DIR)},
        )
    except Exception as e:
        print(f"[run_log] 写入运行记录失败: {e}")

